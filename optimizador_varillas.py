"""
Optimizador de Corte de Varillas
Sistema profesional para minimizar desperdicios en corte de varillas de construcción
"""

import pandas as pd
import numpy as np
from typing import List, Dict, Tuple
from dataclasses import dataclass
from itertools import combinations_with_replacement
import json


@dataclass
class Pieza:
    """Representa una pieza requerida"""
    longitud: float
    diametro: str
    cantidad: int
    elemento: str = ""


@dataclass
class VarillaEstandar:
    """Representa una varilla estándar disponible"""
    longitud: float
    diametro: str
    nombre_referencia: str


@dataclass
class PatronCorte:
    """Representa un patrón de corte optimizado"""
    varilla: VarillaEstandar
    cortes: List[Tuple[float, int]]  # (longitud, cantidad)
    desperdicio: float
    cantidad_varillas: int = 1
    sobrantes_generados: List[float] = None  # Lista de sobrantes que genera este patrón
    sobrantes_usados: List[Tuple[int, float]] = None  # (id_patron_origen, longitud_usada)
    
    def __post_init__(self):
        if self.sobrantes_generados is None:
            self.sobrantes_generados = []
        if self.sobrantes_usados is None:
            self.sobrantes_usados = []


@dataclass
class Sobrante:
    """Representa un sobrante reutilizable"""
    longitud: float
    diametro: str
    varilla_origen: VarillaEstandar
    patron_id: int
    pedido_id_origen: str  # ID único del pedido que generó este sobrante
    usado: bool = False


class OptimizadorCorteVarillas:
    """Optimizador de corte de varillas con minimización de desperdicios"""
    
    # Referencias estándar del mercado
    REFERENCIAS_ESTANDAR = {
        "3/8": [6, 9, 12],
        "1/2": [6, 9, 12],
        "5/8": [6, 9, 12],
        "3/4": [6, 9, 12],
        "7/8": [6, 9, 12],
        "1": [6, 9, 12]
    }
    
    def __init__(self, ruta_excel: str):
        """
        Inicializa el optimizador
        
        Args:
            ruta_excel: Ruta al archivo Excel con los requerimientos
        """
        self.ruta_excel = ruta_excel
        self.piezas_requeridas: List[Pieza] = []
        self.soluciones: Dict[str, List[PatronCorte]] = {}
        self.sobrantes_disponibles: Dict[str, List[Sobrante]] = {}  # Por diámetro
        self.patron_counter = 0  # Contador global para IDs de patrones
        self.pedido_counter = 1  # Contador para IDs únicos de pedidos
        
    def leer_excel(self) -> pd.DataFrame:
        """Lee el archivo Excel y retorna el DataFrame"""
        try:
            df = pd.read_excel(self.ruta_excel)
            print(f"✓ Archivo leído exitosamente: {len(df)} filas")
            print(f"✓ Columnas encontradas: {list(df.columns)}")
            return df
        except Exception as e:
            print(f"✗ Error al leer el archivo: {e}")
            raise
    
    def procesar_datos(self, df: pd.DataFrame):
        """Procesa el DataFrame y extrae las piezas requeridas"""
        self.piezas_requeridas = []
        
        # Agrupar por diámetro y longitud para consolidar cantidades
        df_agrupado = df.groupby(['Ø (in)', 'Length (m)'])['Element Qty'].sum().reset_index()
        
        for _, row in df_agrupado.iterrows():
            # Limpiar el diámetro (remover espacios)
            diametro = str(row['Ø (in)']).strip()
            longitud = float(row['Length (m)'])
            cantidad = int(row['Element Qty'])
            
            if cantidad > 0:
                pieza = Pieza(
                    longitud=longitud,
                    diametro=diametro,
                    cantidad=cantidad
                )
                self.piezas_requeridas.append(pieza)
        
        print(f"\n✓ Piezas procesadas:")
        for p in self.piezas_requeridas:
            print(f"  - {p.cantidad} piezas de {p.longitud}m en diámetro {p.diametro}")
    
    def generar_varillas_disponibles(self, diametro: str) -> List[VarillaEstandar]:
        """Genera lista de varillas estándar disponibles para un diámetro"""
        varillas = []
        if diametro in self.REFERENCIAS_ESTANDAR:
            longitudes = self.REFERENCIAS_ESTANDAR[diametro]
            for i, long in enumerate(longitudes, 1):
                varilla = VarillaEstandar(
                    longitud=long,
                    diametro=diametro,
                    nombre_referencia=f"Ref{i}"
                )
                varillas.append(varilla)
        return varillas
    
    def calcular_patron_simple(self, pieza: Pieza, varilla: VarillaEstandar) -> PatronCorte:
        """Calcula cuántas piezas caben en una varilla y el desperdicio"""
        piezas_por_varilla = int(varilla.longitud // pieza.longitud)
        desperdicio = varilla.longitud - (piezas_por_varilla * pieza.longitud)
        
        return PatronCorte(
            varilla=varilla,
            cortes=[(pieza.longitud, piezas_por_varilla)],
            desperdicio=desperdicio,
            cantidad_varillas=1
        )
    
    def analizar_mejor_referencia_lote(self, pieza: Pieza, varillas_disponibles: list, piezas_pendientes_diametro: list) -> dict:
        """
        Analiza qué referencia (6m, 9m, 12m) es mejor para un lote considerando:
        1. Desperdicio directo
        2. Potencial de reutilización del sobrante para otras piezas del mismo diámetro
        """
        mejor_opcion = None
        mejor_score = float('inf')
        
        for varilla in varillas_disponibles:
            patron = self.calcular_patron_simple(pieza, varilla)
            
            if patron.cortes[0][1] == 0:  # No caben piezas
                continue
            
            piezas_por_varilla = patron.cortes[0][1]
            varillas_necesarias = int(np.ceil(pieza.cantidad / piezas_por_varilla))
            desperdicio_por_varilla = patron.desperdicio
            
            # Calcular desperdicio total directo
            desperdicio_total = desperdicio_por_varilla * varillas_necesarias
            
            # Analizar si el sobrante puede usarse para otras piezas pendientes
            utilidad_sobrante = 0
            if desperdicio_por_varilla >= 0.5:  # Sobrante útil
                for pieza_futura in piezas_pendientes_diametro:
                    if pieza_futura.longitud <= desperdicio_por_varilla:
                        # Este sobrante podría servir para piezas futuras
                        piezas_aprovechables = int(desperdicio_por_varilla / pieza_futura.longitud)
                        utilidad_sobrante += piezas_aprovechables * pieza_futura.longitud
                        break  # Con la primera coincidencia es suficiente para el score
            
            # Score: penaliza desperdicio, premia sobrantes útiles
            score = desperdicio_total - (utilidad_sobrante * varillas_necesarias * 0.8)
            
            if score < mejor_score:
                mejor_score = score
                mejor_opcion = {
                    'varilla': varilla,
                    'patron': patron,
                    'varillas_necesarias': varillas_necesarias,
                    'desperdicio_total': desperdicio_total,
                    'utilidad_sobrante': utilidad_sobrante,
                    'score': score
                }
        
        return mejor_opcion
    
    def buscar_sobrantes_utilizables(self, diametro: str, longitud_necesaria: float) -> List[Sobrante]:
        """Busca sobrantes disponibles que puedan cubrir la longitud necesaria"""
        if diametro not in self.sobrantes_disponibles:
            return []
        
        sobrantes_utiles = [
            s for s in self.sobrantes_disponibles[diametro]
            if not s.usado and s.longitud >= longitud_necesaria
        ]
        
        # Ordenar por longitud (preferir el que más se ajuste)
        sobrantes_utiles.sort(key=lambda x: x.longitud)
        return sobrantes_utiles
    
    def agregar_sobrante(self, diametro: str, longitud: float, varilla: VarillaEstandar, patron_id: int, pedido_id: str):
        """Registra un sobrante para reutilización futura"""
        # Solo registrar sobrantes útiles (mayor a 0.5m)
        if longitud >= 0.5:
            if diametro not in self.sobrantes_disponibles:
                self.sobrantes_disponibles[diametro] = []
            
            sobrante = Sobrante(
                longitud=longitud,
                diametro=diametro,
                varilla_origen=varilla,
                patron_id=patron_id,
                pedido_id_origen=pedido_id
            )
            self.sobrantes_disponibles[diametro].append(sobrante)
    
    def optimizar_con_sobrantes(self, pieza: Pieza, varillas_disponibles: List[VarillaEstandar], pedido_id: str, piezas_pendientes: list = []) -> Dict:
        """Optimiza considerando primero los sobrantes disponibles y el análisis de lote"""
        cantidad_restante = pieza.cantidad
        piezas_desde_sobrantes = 0
        sobrantes_usados_info = []
        
        # Paso 1: Intentar usar sobrantes disponibles
        sobrantes = self.buscar_sobrantes_utilizables(pieza.diametro, pieza.longitud)
        
        for sobrante in sobrantes:
            if cantidad_restante <= 0:
                break
            
            # Calcular cuántas piezas caben en este sobrante
            piezas_en_sobrante = int(sobrante.longitud // pieza.longitud)
            piezas_a_usar = min(piezas_en_sobrante, cantidad_restante)
            
            if piezas_a_usar > 0:
                longitud_usada = piezas_a_usar * pieza.longitud
                sobrante_nuevo = sobrante.longitud - longitud_usada
                
                # Marcar sobrante como usado
                sobrante.usado = True
                
                # Registrar el uso con ID de origen
                sobrantes_usados_info.append({
                    'patron_id': sobrante.patron_id,
                    'pedido_id_origen': sobrante.pedido_id_origen,  # ID del pedido que generó el sobrante
                    'longitud_original': sobrante.longitud,
                    'longitud_usada': longitud_usada,
                    'piezas_obtenidas': piezas_a_usar,
                    'nuevo_sobrante': sobrante_nuevo,
                    'varilla_origen': sobrante.varilla_origen
                })
                
                cantidad_restante -= piezas_a_usar
                piezas_desde_sobrantes += piezas_a_usar
                
                # Si queda un nuevo sobrante útil, registrarlo
                if sobrante_nuevo >= 0.5:
                    self.agregar_sobrante(
                        pieza.diametro,
                        sobrante_nuevo,
                        sobrante.varilla_origen,
                        self.patron_counter,
                        pedido_id  # El pedido actual genera un nuevo sobrante
                    )
        
        # Paso 2: Optimizar para las piezas restantes con varillas nuevas usando análisis de lote
        mejor_solucion = None
        desperdicio_varillas_nuevas = 0
        
        if cantidad_restante > 0:
            # Usar el análisis de mejor referencia considerando el lote
            resultado_lote = self.analizar_mejor_referencia_lote(pieza, varillas_disponibles, piezas_pendientes)
            
            if resultado_lote:
                mejor_patron = resultado_lote['patron']
                piezas_por_varilla = mejor_patron.cortes[0][1]
                varillas_necesarias = int(np.ceil(cantidad_restante / piezas_por_varilla))
                mejor_patron.cantidad_varillas = varillas_necesarias
                mejor_solucion = mejor_patron
                desperdicio_varillas_nuevas = resultado_lote['desperdicio_total']
        
        return {
            'piezas_desde_sobrantes': piezas_desde_sobrantes,
            'sobrantes_usados': sobrantes_usados_info,
            'cantidad_restante': cantidad_restante,
            'patron_varillas_nuevas': mejor_solucion,
            'desperdicio_varillas_nuevas': desperdicio_varillas_nuevas
        }
    
    def optimizar_por_diametro(self, diametro: str):
        """Optimiza el corte para todas las piezas de un diámetro específico"""
        piezas_diametro = [p for p in self.piezas_requeridas if p.diametro == diametro]
        
        if not piezas_diametro:
            return
        
        print(f"\n{'='*60}")
        print(f"OPTIMIZANDO DIÁMETRO {diametro}")
        print(f"{'='*60}")
        
        # Ordenar piezas por longitud (de mayor a menor) para mejor aprovechamiento
        piezas_diametro.sort(key=lambda x: x.longitud, reverse=True)
        
        varillas_disponibles = self.generar_varillas_disponibles(diametro)
        
        if not varillas_disponibles:
            print(f"⚠ No hay referencias estándar para diámetro {diametro}")
            return
        
        soluciones_diametro = []
        
        # Procesar cada pieza pasando las piezas pendientes para optimización de lote
        for idx, pieza in enumerate(piezas_diametro):
            # Generar ID único para este pedido
            pedido_id = f"{diametro}-{self.pedido_counter:03d}"
            self.pedido_counter += 1
            
            # Construir lista de piezas pendientes (las que vienen después)
            piezas_pendientes = piezas_diametro[idx+1:] if idx < len(piezas_diametro) - 1 else []
            
            print(f"\n→ [{pedido_id}] Optimizando: {pieza.cantidad} piezas de {pieza.longitud}m")
            
            # Optimizar con reutilización de sobrantes y análisis de lote
            resultado = self.optimizar_con_sobrantes(pieza, varillas_disponibles, pedido_id, piezas_pendientes)
            
            # Mostrar información de sobrantes usados
            if resultado['piezas_desde_sobrantes'] > 0:
                print(f"  ✓ {resultado['piezas_desde_sobrantes']} piezas obtenidas de SOBRANTES:")
                for info_sobrante in resultado['sobrantes_usados']:
                    print(f"    • Sobrante de pedido [{info_sobrante['pedido_id_origen']}]: {info_sobrante['longitud_original']:.2f}m "
                          f"({info_sobrante['varilla_origen'].nombre_referencia} {info_sobrante['varilla_origen'].longitud}m) → "
                          f"usados {info_sobrante['longitud_usada']:.2f}m → "
                          f"{info_sobrante['piezas_obtenidas']} piezas → "
                          f"queda {info_sobrante['nuevo_sobrante']:.2f}m")
            
            # Mostrar solución para piezas restantes
            if resultado['cantidad_restante'] > 0:
                patron = resultado['patron_varillas_nuevas']
                if patron:
                    varilla = patron.varilla
                    piezas_por_varilla = patron.cortes[0][1]
                    
                    print(f"  • {resultado['cantidad_restante']} piezas desde VARILLAS NUEVAS:")
                    print(f"    {varilla.nombre_referencia} ({varilla.longitud}m): "
                          f"{piezas_por_varilla} piezas/varilla → "
                          f"{patron.cantidad_varillas} varillas → "
                          f"desperdicio: {resultado['desperdicio_varillas_nuevas']:.2f}m")
                    
                    # Registrar sobrantes generados por las varillas nuevas
                    if patron.desperdicio > 0:
                        for _ in range(patron.cantidad_varillas):
                            self.agregar_sobrante(
                                diametro,
                                patron.desperdicio,
                                varilla,
                                self.patron_counter,
                                pedido_id
                            )
                    
                    self.patron_counter += 1
                    
                    soluciones_diametro.append({
                        'pieza': pieza,
                        'patron': patron,
                        'sobrantes_usados': resultado['sobrantes_usados'],
                        'piezas_desde_sobrantes': resultado['piezas_desde_sobrantes'],
                        'desperdicio_total': resultado['desperdicio_varillas_nuevas'],
                        'pedido_id': pedido_id
                    })
            else:
                # Todas las piezas se obtuvieron de sobrantes
                print(f"  ✓ TODAS las piezas obtenidas de sobrantes - ¡0 varillas nuevas!")
                soluciones_diametro.append({
                    'pieza': pieza,
                    'patron': None,
                    'sobrantes_usados': resultado['sobrantes_usados'],
                    'piezas_desde_sobrantes': resultado['piezas_desde_sobrantes'],
                    'desperdicio_total': 0,
                    'pedido_id': pedido_id
                })
        
        self.soluciones[diametro] = soluciones_diametro
    
    def optimizar(self):
        """Ejecuta la optimización completa"""
        print("\n" + "="*60)
        print("INICIANDO OPTIMIZACIÓN DE CORTE DE VARILLAS")
        print("="*60)
        
        # Leer y procesar datos
        df = self.leer_excel()
        self.procesar_datos(df)
        
        # Obtener diámetros únicos
        diametros_unicos = list(set(p.diametro for p in self.piezas_requeridas))
        
        # Optimizar por cada diámetro
        for diametro in diametros_unicos:
            self.optimizar_por_diametro(diametro)
    
    def generar_reporte(self) -> str:
        """Genera un reporte detallado de la solución óptima"""
        reporte = []
        reporte.append("\n" + "="*80)
        reporte.append("REPORTE DE OPTIMIZACIÓN - PLAN DE CORTE Y COMPRA CON REUTILIZACIÓN")
        reporte.append("="*80)
        
        desperdicio_total_general = 0
        desperdicio_real_final = 0
        total_piezas_desde_sobrantes = 0
        
        for diametro, soluciones in self.soluciones.items():
            reporte.append(f"\n{'─'*80}")
            reporte.append(f"DIÁMETRO: {diametro}\"")
            reporte.append(f"{'─'*80}")
            
            # Resumen de compra por referencia
            compra_por_ref = {}
            
            for sol in soluciones:
                pieza = sol['pieza']
                patron = sol['patron']
                desperdicio = sol['desperdicio_total']
                piezas_sobrantes = sol['piezas_desde_sobrantes']
                sobrantes_usados = sol['sobrantes_usados']
                pedido_id = sol['pedido_id']
                
                reporte.append(f"\n📏 [{pedido_id}] Pieza requerida: {pieza.longitud}m × {pieza.cantidad} unidades")
                
                # Mostrar uso de sobrantes
                if piezas_sobrantes > 0:
                    reporte.append(f"   ✓ {piezas_sobrantes} piezas obtenidas de SOBRANTES:")
                    for info in sobrantes_usados:
                        reporte.append(f"      • DE PEDIDO [{info['pedido_id_origen']}]: {info['longitud_usada']:.2f}m de sobrante "
                                     f"({info['varilla_origen'].nombre_referencia} {info['varilla_origen'].longitud}m) "
                                     f"→ {info['piezas_obtenidas']} piezas")
                    total_piezas_desde_sobrantes += piezas_sobrantes
                
                # Mostrar varillas nuevas necesarias
                if patron:
                    varilla = patron.varilla
                    ref_key = f"{varilla.nombre_referencia} ({varilla.longitud}m)"
                    
                    if ref_key not in compra_por_ref:
                        compra_por_ref[ref_key] = 0
                    compra_por_ref[ref_key] += patron.cantidad_varillas
                    
                    piezas_nuevas = pieza.cantidad - piezas_sobrantes
                    reporte.append(f"   Varillas nuevas: {patron.cantidad_varillas} × {ref_key}")
                    reporte.append(f"   → {patron.cortes[0][1]} piezas por varilla")
                    reporte.append(f"   → Produce {piezas_nuevas} piezas")
                    reporte.append(f"   → Desperdicio: {desperdicio:.2f}m")
                    
                    desperdicio_total_general += desperdicio
                else:
                    reporte.append(f"   ✅ 0 varillas nuevas necesarias (todo desde sobrantes)")
            
            reporte.append(f"\n{'─'*40}")
            reporte.append(f"LISTA DE COMPRA - DIÁMETRO {diametro}\":")
            reporte.append(f"{'─'*40}")
            if compra_por_ref:
                for ref, cantidad in sorted(compra_por_ref.items()):
                    reporte.append(f"  🛒 {cantidad} varillas de {ref}")
            else:
                reporte.append(f"  ✅ No se necesitan varillas nuevas (todo desde sobrantes)")
        
        # Calcular sobrantes finales no utilizados
        for diametro, sobrantes in self.sobrantes_disponibles.items():
            for sobrante in sobrantes:
                if not sobrante.usado:
                    desperdicio_real_final += sobrante.longitud
        
        reporte.append(f"\n{'='*80}")
        reporte.append(f"RESUMEN GENERAL")
        reporte.append(f"{'='*80}")
        reporte.append(f"📊 Desperdicio en cortes: {desperdicio_total_general:.2f} metros")
        reporte.append(f"📊 Sobrantes finales no usados: {desperdicio_real_final:.2f} metros")
        reporte.append(f"📊 Desperdicio REAL total: {desperdicio_total_general + desperdicio_real_final:.2f} metros")
        reporte.append(f"♻️  Total de piezas obtenidas de sobrantes: {total_piezas_desde_sobrantes}")
        reporte.append(f"✅ Optimización con reutilización completada exitosamente")
        reporte.append(f"{'='*80}\n")
        
        return "\n".join(reporte)
    
    def exportar_plan_corte(self, ruta_salida: str = None):
        """Exporta el plan de corte detallado a Excel"""
        if ruta_salida is None:
            ruta_salida = self.ruta_excel.replace('.xlsx', '_PLAN_CORTE_OPTIMIZADO.xlsx')
        
        datos_export = []
        
        for diametro, soluciones in self.soluciones.items():
            for sol in soluciones:
                pieza = sol['pieza']
                patron = sol['patron']
                desperdicio = sol['desperdicio_total']
                piezas_sobrantes = sol['piezas_desde_sobrantes']
                sobrantes_usados = sol['sobrantes_usados']
                pedido_id = sol['pedido_id']
                
                # Información de sobrantes usados con IDs
                info_sobrantes = ""
                ids_origen_sobrantes = ""
                if piezas_sobrantes > 0:
                    detalles = [f"{s['piezas_obtenidas']}pz de [{s['pedido_id_origen']}]" 
                               for s in sobrantes_usados]
                    info_sobrantes = "; ".join(detalles)
                    
                    # IDs únicos de los pedidos origen
                    ids_unicos = list(set([s['pedido_id_origen'] for s in sobrantes_usados]))
                    ids_origen_sobrantes = ", ".join(ids_unicos)
                
                # Varillas nuevas
                varillas_nuevas = 0
                ref_varilla = "N/A"
                piezas_por_varilla = 0
                sobrante_generado = 0.0
                
                if patron:
                    varillas_nuevas = patron.cantidad_varillas
                    ref_varilla = f"{patron.varilla.nombre_referencia} ({patron.varilla.longitud}m)"
                    piezas_por_varilla = patron.cortes[0][1]
                    sobrante_generado = patron.desperdicio
                
                datos_export.append({
                    'ID Pedido': pedido_id,
                    'Diámetro': diametro,
                    'Longitud Pieza (m)': pieza.longitud,
                    'Cantidad Total Piezas': pieza.cantidad,
                    'Piezas de Sobrantes': piezas_sobrantes,
                    'Sobrantes de Pedidos': ids_origen_sobrantes,
                    'Detalle Sobrantes Usados': info_sobrantes,
                    'Varillas Nuevas a Comprar': varillas_nuevas,
                    'Referencia Varilla': ref_varilla,
                    'Piezas por Varilla Nueva': piezas_por_varilla if patron else 0,
                    'Sobrante por Varilla (m)': sobrante_generado,
                    'Sobrante Total Generado (m)': round(sobrante_generado * varillas_nuevas, 2) if patron else 0,
                    'Desperdicio (m)': round(desperdicio, 2),
                    'Eficiencia %': round((1 - desperdicio / (patron.varilla.longitud * varillas_nuevas)) * 100, 2) if patron and varillas_nuevas > 0 else 100
                })
        
        df_export = pd.DataFrame(datos_export)
        
        # Crear un escritor de Excel con múltiples hojas
        with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
            df_export.to_excel(writer, index=False, sheet_name='Plan de Corte')
            
            # Hoja de resumen
            resumen_data = []
            for diametro, soluciones in self.soluciones.items():
                compra_por_ref = {}
                total_piezas_sobrantes = 0
                
                for sol in soluciones:
                    patron = sol['patron']
                    total_piezas_sobrantes += sol['piezas_desde_sobrantes']
                    
                    if patron:
                        ref_key = f"{patron.varilla.nombre_referencia} ({patron.varilla.longitud}m)"
                        if ref_key not in compra_por_ref:
                            compra_por_ref[ref_key] = 0
                        compra_por_ref[ref_key] += patron.cantidad_varillas
                
                for ref, cantidad in compra_por_ref.items():
                    resumen_data.append({
                        'Diámetro': diametro,
                        'Referencia': ref,
                        'Cantidad a Comprar': cantidad,
                        'Piezas desde Sobrantes': total_piezas_sobrantes
                    })
            
            df_resumen = pd.DataFrame(resumen_data)
            df_resumen.to_excel(writer, index=False, sheet_name='Lista de Compra')
            
            # Hoja de trazabilidad de sobrantes
            trazabilidad_data = []
            for diametro, soluciones in self.soluciones.items():
                for sol in soluciones:
                    pedido_id = sol['pedido_id']
                    pieza = sol['pieza']
                    sobrantes_usados = sol['sobrantes_usados']
                    
                    if sobrantes_usados:
                        for info in sobrantes_usados:
                            trazabilidad_data.append({
                                'Pedido Destino': pedido_id,
                                'Pedido Origen Sobrante': info['pedido_id_origen'],
                                'Diámetro': diametro,
                                'Longitud Sobrante Original (m)': info['longitud_original'],
                                'Longitud Usada (m)': info['longitud_usada'],
                                'Piezas Obtenidas': info['piezas_obtenidas'],
                                'Longitud Pieza (m)': pieza.longitud,
                                'Varilla Origen': f"{info['varilla_origen'].nombre_referencia} ({info['varilla_origen'].longitud}m)"
                            })
            
            if trazabilidad_data:
                df_trazabilidad = pd.DataFrame(trazabilidad_data)
                df_trazabilidad.to_excel(writer, index=False, sheet_name='Trazabilidad Sobrantes')
        
        print(f"\n✅ Plan de corte optimizado exportado a: {ruta_salida}")
        return ruta_salida
    
    def generar_orden_compra(self, ruta_salida: str = None):
        """Genera una orden de compra totalizada lista para el proveedor"""
        if ruta_salida is None:
            ruta_salida = self.ruta_excel.replace('.xlsx', '_ORDEN_COMPRA.xlsx')
        
        # Consolidar todas las varillas por diámetro y referencia
        orden_compra = {}
        
        for diametro, soluciones in self.soluciones.items():
            if diametro not in orden_compra:
                orden_compra[diametro] = {}
            
            for sol in soluciones:
                patron = sol['patron']
                if patron:
                    ref_key = f"{patron.varilla.nombre_referencia}"
                    longitud = patron.varilla.longitud
                    
                    if ref_key not in orden_compra[diametro]:
                        orden_compra[diametro][ref_key] = {
                            'longitud': longitud,
                            'cantidad': 0
                        }
                    
                    orden_compra[diametro][ref_key]['cantidad'] += patron.cantidad_varillas
        
        # Crear DataFrame para la orden de compra
        datos_orden = []
        item_numero = 1
        
        for diametro in sorted(orden_compra.keys()):
            refs = orden_compra[diametro]
            for ref in sorted(refs.keys(), key=lambda x: refs[x]['longitud']):
                datos_orden.append({
                    'Item': item_numero,
                    'Diámetro (in)': diametro,
                    'Referencia': ref,
                    'Longitud (m)': refs[ref]['longitud'],
                    'Cantidad': refs[ref]['cantidad'],
                    'Unidad': 'Varillas',
                    'Observaciones': f"Varillas de {diametro}\" × {refs[ref]['longitud']}m"
                })
                item_numero += 1
        
        df_orden = pd.DataFrame(datos_orden)
        
        # Calcular totales
        total_varillas = df_orden['Cantidad'].sum()
        total_items = len(df_orden)
        
        # Crear resumen por diámetro
        resumen_diametro = []
        for diametro in sorted(orden_compra.keys()):
            refs = orden_compra[diametro]
            total_por_diametro = sum(refs[ref]['cantidad'] for ref in refs)
            resumen_diametro.append({
                'Diámetro': diametro,
                'Total Varillas': total_por_diametro
            })
        
        df_resumen_diametro = pd.DataFrame(resumen_diametro)
        
        # Exportar a Excel con formato profesional
        with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
            # Hoja principal - Orden de Compra
            df_orden.to_excel(writer, index=False, sheet_name='Orden de Compra', startrow=4)
            
            # Obtener el workbook y worksheet
            workbook = writer.book
            worksheet = writer.sheets['Orden de Compra']
            
            # Agregar encabezado
            from datetime import datetime
            worksheet['A1'] = 'ORDEN DE COMPRA - VARILLAS DE CONSTRUCCIÓN'
            worksheet['A2'] = f'Fecha: {datetime.now().strftime("%d/%m/%Y")}'
            worksheet['A3'] = f'Total Items: {total_items} | Total Varillas: {total_varillas}'
            
            # Formato de encabezado
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            # Título principal
            worksheet['A1'].font = Font(size=14, bold=True)
            worksheet['A1'].alignment = Alignment(horizontal='center')
            worksheet.merge_cells('A1:G1')
            
            # Subtítulos
            worksheet['A2'].font = Font(size=11)
            worksheet['A3'].font = Font(size=11, bold=True)
            
            # Formatear tabla
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Aplicar formato a encabezados
            for col in range(1, 8):
                cell = worksheet.cell(row=5, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Aplicar bordes a datos
            for row in range(6, 6 + len(df_orden)):
                for col in range(1, 8):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Ajustar ancho de columnas
            worksheet.column_dimensions['A'].width = 8
            worksheet.column_dimensions['B'].width = 15
            worksheet.column_dimensions['C'].width = 12
            worksheet.column_dimensions['D'].width = 15
            worksheet.column_dimensions['E'].width = 12
            worksheet.column_dimensions['F'].width = 12
            worksheet.column_dimensions['G'].width = 35
            
            # Agregar fila de totales
            total_row = 6 + len(df_orden)
            worksheet.cell(row=total_row, column=4).value = 'TOTAL:'
            worksheet.cell(row=total_row, column=4).font = Font(bold=True)
            worksheet.cell(row=total_row, column=5).value = total_varillas
            worksheet.cell(row=total_row, column=5).font = Font(bold=True)
            worksheet.cell(row=total_row, column=6).value = 'Varillas'
            worksheet.cell(row=total_row, column=6).font = Font(bold=True)
            
            # Hoja de resumen por diámetro
            df_resumen_diametro.to_excel(writer, index=False, sheet_name='Resumen por Diámetro')
            worksheet_resumen = writer.sheets['Resumen por Diámetro']
            
            # Formatear resumen
            for col in range(1, 3):
                cell = worksheet_resumen.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            for row in range(2, 2 + len(df_resumen_diametro)):
                for col in range(1, 3):
                    cell = worksheet_resumen.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            worksheet_resumen.column_dimensions['A'].width = 15
            worksheet_resumen.column_dimensions['B'].width = 15
            
            # Agregar total general en resumen
            total_row_resumen = 2 + len(df_resumen_diametro)
            worksheet_resumen.cell(row=total_row_resumen, column=1).value = 'TOTAL GENERAL'
            worksheet_resumen.cell(row=total_row_resumen, column=1).font = Font(bold=True)
            worksheet_resumen.cell(row=total_row_resumen, column=2).value = total_varillas
            worksheet_resumen.cell(row=total_row_resumen, column=2).font = Font(bold=True)
        
        print(f"\n✅ Orden de compra generada: {ruta_salida}")
        print(f"   📦 Total de items: {total_items}")
        print(f"   📊 Total de varillas: {total_varillas}")
        
        return ruta_salida


def main():
    """Función principal"""
    import os
    
    # Ruta al archivo Excel
    ruta_excel = os.path.join(os.environ['USERPROFILE'], 'Downloads', 'Cortes.xlsx')
    
    # Crear optimizador
    optimizador = OptimizadorCorteVarillas(ruta_excel)
    
    # Ejecutar optimización
    optimizador.optimizar()
    
    # Generar y mostrar reporte
    reporte = optimizador.generar_reporte()
    print(reporte)
    
    # Exportar plan de corte
    optimizador.exportar_plan_corte()
    
    # Generar orden de compra
    optimizador.generar_orden_compra()


if __name__ == "__main__":
    main()
